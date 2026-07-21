import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class ExcelSorterNeo:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Sorter Neo 🔮")
        self.root.geometry("640x500")
        self.root.configure(bg="#0e0f11")

        self.archivo = None
        self.sheetnames = []
        self.df = None

        # Estilos visuales personalizados
        self._custom_styles()

        # Contenedor principal estilo tarjeta
        self.card = tk.Frame(root, bg="#1a1d24", bd=0, highlightbackground="#2e3540", highlightthickness=2)
        self.card.place(relx=0.5, rely=0.5, anchor="center", width=520, height=430)

        # Título
        tk.Label(self.card, text="🔠 Excel Sorter Neo", font=("Segoe UI Semibold", 20),
                 fg="#00f0ff", bg="#1a1d24").pack(pady=(25, 10))

        # Botón seleccionar archivo
        ttk.Button(self.card, text="📂 Seleccionar archivo", command=self.cargar_excel, style="Neo.TButton")\
            .pack(pady=10, ipadx=10)

        # Controles dinámicos
        self.hoja_label = tk.Label(self.card, text="Selecciona la hoja:", font=("Segoe UI", 11), fg="#ffffff", bg="#1a1d24")
        self.hoja_combo = ttk.Combobox(self.card, state='readonly', style="Neo.TCombobox")

        self.columna_label = tk.Label(self.card, text="Selecciona la columna de nombres:", font=("Segoe UI", 11), fg="#ffffff", bg="#1a1d24")
        self.columna_combo = ttk.Combobox(self.card, state='readonly', style="Neo.TCombobox")

        self.orden_label = tk.Label(self.card, text="Orden:", font=("Segoe UI", 11), fg="#ffffff", bg="#1a1d24")
        self.orden_combo = ttk.Combobox(self.card, state='readonly', style="Neo.TCombobox", values=["Ascendente", "Descendente"])
        self.orden_combo.current(0)  # Por defecto, ascendente

        self.ordenar_btn = ttk.Button(self.card, text="✅ Ordenar y guardar", command=self.ordenar_y_guardar, style="Neo.TButton")

        # Barra de progreso
        self.progress = ttk.Progressbar(self.card, style="Neo.Horizontal.TProgressbar", mode='indeterminate')

    def _custom_styles(self):
        style = ttk.Style()
        style.theme_use("clam")

        # Botón
        style.configure("Neo.TButton",
                        font=("Segoe UI", 11, "bold"),
                        padding=10,
                        foreground="#00f0ff",
                        background="#2b2f38",
                        relief="flat",
                        borderwidth=0)
        style.map("Neo.TButton",
                  background=[("active", "#00f0ff")],
                  foreground=[("active", "#0e0f11")])

        # Combobox
        style.configure("Neo.TCombobox",
                        fieldbackground="#2b2f38",
                        background="#2b2f38",
                        foreground="#00f0ff",
                        arrowcolor="#00f0ff",
                        bordercolor="#00f0ff")
        style.map("Neo.TCombobox",
                  fieldbackground=[("readonly", "#2b2f38")],
                  background=[("readonly", "#2b2f38")],
                  foreground=[("readonly", "#00f0ff")])

        # Barra de progreso
        style.configure("Neo.Horizontal.TProgressbar",
                        troughcolor="#2b2f38",
                        background="#00f0ff",
                        bordercolor="#00f0ff")

    def cargar_excel(self):
        archivo = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if not archivo:
            return
        try:
            self.archivo = archivo
            self.sheetnames = pd.ExcelFile(self.archivo).sheet_names
            if not self.sheetnames:
                raise ValueError("El archivo no contiene hojas válidas.")
            self.hoja_combo['values'] = self.sheetnames
            self.hoja_combo.current(0)

            self.hoja_label.pack(pady=(20, 5))
            self.hoja_combo.pack()
            self.orden_label.pack(pady=(10, 5))
            self.orden_combo.pack()

            messagebox.showinfo("✔ Archivo cargado", f"{archivo.split('/')[-1]}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo cargar el archivo: {str(e)}")

        self.hoja_combo.bind("<<ComboboxSelected>>", self.cargar_hoja)

    def cargar_hoja(self, event=None):
        hoja = self.hoja_combo.get()
        try:
            self.df = pd.read_excel(self.archivo, sheet_name=hoja)
            if self.df.empty:
                raise ValueError("La hoja seleccionada está vacía.")
            if len(self.df.columns) < 2:
                raise ValueError("El archivo debe tener al menos dos columnas: numeración y nombres.")

            # Validar que la primera columna (A) contenga numeración
            first_column = self.df.columns[0]
            if not pd.api.types.is_numeric_dtype(self.df[first_column]):
                raise ValueError("La primera columna (A) debe contener valores numéricos (numeración).")

            # Mostrar solo las columnas a partir de la segunda para ordenar (nombres)
            self.columna_combo['values'] = list(self.df.columns[1:])
            self.columna_combo.current(0)

            self.columna_label.pack(pady=(20, 5))
            self.columna_combo.pack()
            self.ordenar_btn.pack(pady=30, ipadx=5)
        except Exception as e:
            messagebox.showerror("Error al leer la hoja", str(e))

    def ordenar_y_guardar(self):
        columna = self.columna_combo.get()
        if not columna:
            messagebox.showwarning("⚠", "Selecciona una columna de nombres.")
            return
        try:
            # Iniciar barra de progreso
            self.progress.pack(pady=10)
            self.progress.start()
            self.root.update()

            # Validar que la columna seleccionada contenga texto
            if not pd.api.types.is_string_dtype(self.df[columna]):
                raise ValueError("La columna seleccionada debe contener nombres (texto).")

            # Determinar tipo de orden
            ascending = self.orden_combo.get() == "Ascendente"
            df_ordenado = self.df.sort_values(by=columna, ascending=ascending)

            # Diálogo de guardar archivo
            nuevo_archivo = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                         filetypes=[("Excel files", "*.xlsx")])
            if not nuevo_archivo:
                self.progress.stop()
                self.progress.pack_forget()
                return

            # Guardar el archivo inicialmente
            df_ordenado.to_excel(nuevo_archivo, index=False)

            # Estilo para el archivo
            wb = load_workbook(nuevo_archivo)
            ws = wb.active

            # Estilo para encabezado
            header_fill = PatternFill(start_color="007acc", end_color="007acc", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, name='Segoe UI')
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            # Autoajuste de columnas con ancho mínimo
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = max((max_length + 2), 10)  # Ancho mínimo de 10
                ws.column_dimensions[column].width = adjusted_width

            # Estilo para celdas normales
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = Font(name='Segoe UI', size=10)
                    # Centrar si la celda pertenece a la columna A (numeración) o es numérica
                    alignment = Alignment(horizontal="center" if cell.column == 1 or isinstance(cell.value, (int, float)) else "left",
                                         vertical="center")
                    cell.alignment = alignment
                    cell.border = thin_border

            wb.save(nuevo_archivo)

            # Detener barra de progreso
            self.progress.stop()
            self.progress.pack_forget()

            messagebox.showinfo("✅ Excel Guardado", f"Archivo ordenado por '{columna}' con numeración sincronizada.")

            # Reiniciar interfaz
            self._reset_ui()
        except Exception as e:
            self.progress.stop()
            self.progress.pack_forget()
            messagebox.showerror("Error", f"No se pudo procesar el archivo: {str(e)}")

    def _reset_ui(self):
        """Reinicia la interfaz para permitir una nueva operación."""
        self.hoja_combo.pack_forget()
        self.columna_combo.pack_forget()
        self.ordenar_btn.pack_forget()
        self.hoja_label.pack_forget()
        self.columna_label.pack_forget()
        self.orden_label.pack_forget()
        self.orden_combo.pack_forget()
        self.archivo = None
        self.sheetnames = []
        self.df = None
        self.hoja_combo['values'] = []
        self.columna_combo['values'] = []

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelSorterNeo(root)
    root.mainloop()